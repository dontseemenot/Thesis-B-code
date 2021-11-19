# Helper functions for algorithm
import numpy as np
from sklearn.utils import resample
import matplotlib.pyplot as plt
import pandas as pd
import os
from sklearn.preprocessing import MinMaxScaler
import openpyxl as pxl
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, accuracy_score
import argparse
from datetime import datetime
from sklearn.model_selection import GroupShuffleSplit
import tensorflow as tf
from tensorflow.keras import layers
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.datasets import fetch_openml
from imblearn.over_sampling import RandomOverSampler
import gc
'''
PREPROCESSING RELATED HELPERS
'''
class ReshapeToTensor():
    def __init__(self):
        pass
    def fit(self, X, y = None):
        return self
    
    def transform(self, X, y = None):
        X = np.expand_dims(X, -1)
        #y = np.expand_dims(y, -1)
        return X
class scale1D():
    def __init__(self):
        pass
    def fit(self, X, y = None):
        return self
    def transform(self, X, y = None):
        X = (X - (-250e-6)) / (250e-6 - (-250e-6))
        assert(np.max(X) <= 1.1)
        assert(np.min(X) >= -0.1)
        return X

class scale2D():
    def __init__(self):
        pass
    def fit(self, X, y = None):
        return self
    def transform(self, X, y = None):
        print(X.dtype)
        X = X.astype(np.float32)
        X = X/256.0
        print(np.max(X))
        print(np.min(X))
        # assert(np.max(X) <= 1.0)
        # assert(np.min(X) >= 0.0)
        return X

# A hack to plot confusion matrix with keras model
class MyModelPredict(object):
    def __init__(self, model):
        self._estimator_type = 'classifier'
        self.model = model
        
    def predict(self, X_test):
        m = self.model
        y_pred = m.predict_classes(X_test)
        return y_pred

class Reshape2Dto1D(BaseEstimator, TransformerMixin):
    def __init__(self):
        pass
    def fit(self, X, y = None):
        return self
    def transform(self, X, y = None):
        # print(f"old shape {X.shape}")
        X = np.reshape(X, (X.shape[0], X.shape[1]*X.shape[2]))
        # print(f"new shape {X.shape}")
        return X

class Reshape1Dto2D(BaseEstimator, TransformerMixin):
    def __init__(self, height=227, width=256):
        self.height = height
        self.width = width

    def fit(self, X, y = None):
        return self
    def transform(self, X, y = None):   # Rows, then cols
        # print(f"old shape {X.shape}")
        X = np.reshape(X, (X.shape[0], self.height, self.width, X.shape[2]))
        # print(f"new shape {X.shape}")
        return X

# Mimic StratifiedGroupKFold behaviour
# i.e. achieve as even as possible of class distribution while preserving groups
def group_greedy(df_metadata, pIDs, n_splits):
    group_dict = {}
    group_count_ins = {x: 0 for x in range(n_splits)}
    group_count_con = {x: 0 for x in range(n_splits)}

    df_sorted = df_metadata.sort_values(by = ['Total'], ascending = False)
    for pID in pIDs:
        if df_sorted.loc[df_metadata["pID"] == pID, 'pClass'].values[0] == 'I':
            key = min(group_count_ins, key = group_count_ins.get)
            group_count_ins[key] += df_sorted.loc[df_metadata["pID"] == pID, 'Total'].values[0]
        elif df_sorted.loc[df_metadata["pID"] == pID, 'pClass'].values[0] == 'G':
            key = min(group_count_con, key = group_count_con.get)
            group_count_con[key] += df_sorted.loc[df_metadata["pID"] == pID, 'Total'].values[0] 
        group_dict[pID] = key
        # print(f"ins {group_count_ins} con {group_count_con}")
        # print(f"pID {pID} assigned group {key}")
    return group_dict

# Calculate threshold for number of epochs
def max_num_epochs(stages):
    max_epochs = 800
    num_ALL = stages['W'] + stages['S1'] + stages['S2'] + stages['S3'] + stages['S4'] + stages['R']
    num_LSS = stages['S1'] + stages['S2']
    num_SWS = stages['S3'] + stages['S4']
    num_REM = stages['R']
    num_BSL = stages['W'] + stages['S1'] + stages['S2'] + stages['R']
    threshold = min(max_epochs, num_ALL, num_LSS, num_SWS, num_REM, num_BSL)
    # print(f"W {stages['W']} S1 {stages['S1']} S2 {stages['S2']} S3 {stages['S3']} S4 {stages['S4']} R {stages['R']}")
    # print(f'ALL {num_ALL}, LSS {num_LSS}, SWS {num_SWS}, REM {num_REM}, BSL {num_BSL}\nChosen threshold {threshold}')
    return threshold


def get_sleep_epochs(df, subdataset):
    if subdataset == 'ALL':
        data = df.loc[(df['Sleep_Stage'] == 'W') | (df['Sleep_Stage'] == 'S1') | (df['Sleep_Stage'] == 'S2') | (df['Sleep_Stage'] == 'S3') | (df['Sleep_Stage'] == 'S4') | (df['Sleep_Stage'] == 'R') ]
    elif subdataset == 'LSS':
        data = df.loc[(df['Sleep_Stage'] == 'S1') |  (df['Sleep_Stage'] == 'S2')]
    elif subdataset == 'N1':
        data = df.loc[(df['Sleep_Stage'] == 'S1')]
    elif subdataset == 'N2':
        data = df.loc[(df['Sleep_Stage'] == 'S2')]
    elif subdataset == 'SWS':
        data = df.loc[(df['Sleep_Stage'] == 'S3') |  (df['Sleep_Stage'] == 'S4')]
    elif subdataset == 'REM':
        data = df.loc[(df['Sleep_Stage'] == 'R')]
    elif subdataset == 'BSL':
        data = df.loc[(df['Sleep_Stage'] == 'W') |  (df['Sleep_Stage'] == 'S1') | (df['Sleep_Stage'] == 'S2') | (df['Sleep_Stage'] == 'R')]
    else:
        print("Invalid subdataset specified!")
        data = None
    return data

def balance_dataset(df, threshold, subdataset):
    data = get_sleep_epochs(df, subdataset)

    #print(f'data len: {len(data)}')
    data_resampled = resample(data, replace = False, n_samples = threshold, random_state = 42)
    count = data_resampled['Sleep_Stage'].value_counts()
    # print(f'After balancing:\n{count}')
    assert(len(data_resampled) == threshold)
    return data_resampled

def class_balance(X, y, groups, n_splits):
    X_bal = []
    y_bal = []
    groups_bal = []
    ros = RandomOverSampler(random_state=42, sampling_strategy = 'minority')
    for group in range(n_splits):
        print(f"balancing group {group}")
        group_index = np.where(groups == group)
        X_group = X[group_index]
        y_group = y[group_index]
        print(np.max(X_group))
        print(np.min(X_group))
        print(f"x shape {X_group.shape}")
        print(f"y shape {y_group.shape}")
        X_res, y_res = ros.fit_resample(X_group, y_group)
        for a, b in zip(X_res, y_res):
            X_bal.append(a)
            y_bal.append(b)
            groups_bal.append(group)
        # print(f"{np.bincount(y_res)}")

    return list_to_array(X_bal, y_bal, groups_bal)

def list_to_array(a, b, c):
    a = np.asarray(a)
    b = np.asarray(b)
    c = np.asarray(c)
    return a, b, c

def get_train_test2(X, y, groups, train_index, test_index):
    X_train = X[train_index].astype('float32')
    y_train = y[train_index]
    X_test = X[test_index].astype('float32')
    y_test = y[test_index]
    groups_train = groups[train_index]
    groups_test = groups[test_index]

    return X_train, y_train, groups_train, X_test, y_test, groups_test

'''
SAVING AND CALCULATING RESULTS FUNCTIONS
'''
def create_dirs(title, dt_string):

    results_dir = f'./results/{title} {dt_string}/'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
    models_dir = f'{results_dir}/models'
    if not os.path.exists(models_dir):
        os.makedirs(models_dir)
    images_dir = f'{results_dir}/images'
    if not os.path.exists(images_dir):
        os.makedirs(images_dir)
    return results_dir, models_dir, images_dir

# Initial params
def save_parameters(args, data_info, spreadsheet_file, sheet_name):
    book = pxl.load_workbook(spreadsheet_file)
    writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

    # writer = pd.ExcelWriter(spreadsheet_file, engine='xlsxwriter')   
    # workbook=writer.book
    # worksheet=workbook.add_worksheet(sheet_name)
    # writer.sheets[sheet_name] = worksheet
    offset = 0
    df_args = pd.DataFrame(data = args, index = [0])
    df_args.to_excel(writer, sheet_name = sheet_name, startrow = 1 , startcol = 0, index = False)

    offset = len(df_args) + 3
    df_data_info = pd.DataFrame(data_info, index = [0])
    df_data_info.to_excel(writer, sheet_name = sheet_name, startrow = offset, startcol = 0, index = False)
    offset += len(df_args) + 3
    writer.save()
    return offset

# Mean results saved at the end of program
def save_mean_results(performance_metrics_mean, timestamps, spreadsheet_file, sheet_name, offset):
    book = pxl.load_workbook(spreadsheet_file)
    writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

    df_mean_results = pd.DataFrame(data = performance_metrics_mean, index = [0])
    df_mean_results.to_excel(writer, sheet_name = sheet_name, startrow = offset, startcol = 0, index= False)
    offset += len(df_mean_results) + 3

    df_timestamps = pd.DataFrame(data = timestamps, index = [0])
    df_timestamps.to_excel(writer, sheet_name = sheet_name, startrow = offset, startcol = 0, index= False)
    offset += len(df_timestamps) + 3

    writer.save()
    return offset

# Fold results
def save_fold_results(epoch_counts, performance_metrics, fold_num, spreadsheet_file, sheet_name):
    book = pxl.load_workbook(spreadsheet_file)
    writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

    info = {**epoch_counts, **performance_metrics}

    header = True if fold_num == 0 else None
    offset = 0 if fold_num == 0 else 1
    df_performance = pd.DataFrame(data = info, index = [0])
    df_performance.to_excel(writer, sheet_name = sheet_name, startrow = offset + fold_num, startcol = 0, index= False, header = header)
    writer.save()

def append_summary(args, performance_metrics_mean, timestamps, number, spreadsheet_file, sheet_name):
    book = pxl.load_workbook(spreadsheet_file)
    writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

    data = {**args, **performance_metrics_mean, **timestamps}
    df = pd.DataFrame(data = data, index = [0])
    df.to_excel(writer, sheet_name = sheet_name,startrow = number + 1, startcol = 0, header = False, index = False)


    # df_performance.to_excel(writer, sheet_name = sheet_name, startrow = number + 1 , startcol = 18, index= False, header = False)

    # df_number = pd.DataFrame(data = [number])
    # df_number.to_excel(writer, sheet_name = sheet_name, startrow = number + 1, startcol = 1, index = False, header = False)
    writer.save()

def calculate_performance_metrics(y_test, y_pred, cm, fold_num):
    performance_metrics = {}
    #performance_metrics['fold'] = fold_num
    performance_metrics['accuracy'] = accuracy_score(y_test, y_pred)
    tn, fp, fn, tp = cm.ravel()
    performance_metrics['precision'] = tp/(tp + fp)
    performance_metrics['recall'] = tp/(tp + fn)
    performance_metrics['sensitivity'] = tp/(tp + fn)
    performance_metrics['specificity'] = tn/(tn + fp)
    performance_metrics['f1'] = 2 * performance_metrics['precision'] * performance_metrics['recall'] / (performance_metrics['precision'] + performance_metrics['recall'])
    return performance_metrics

'''
PLOTTING GRAPH FUNCTIONS
'''
def plot_train_val_acc_loss2(train_val_acc_loss, fold_num, images_dir):
    # Accuracy
    plt.rcParams["figure.figsize"] = (10,5)
    plt.plot(train_val_acc_loss['train_acc'], label = 'Training accuracy', color = 'darkorange')
    plt.plot(train_val_acc_loss['valid_acc'], label = 'Validation accuracy', color = 'darkgreen')
    plt.title(f'Best model Accuracy')
    plt.ylabel('Accuracy')
    plt.xlabel('Training epoch')
    plt.yticks(np.linspace(0, 1, num=21))
    plt.grid()
    plt.legend(loc=4)   # lower right
    plt.savefig(f'{images_dir}/Fold {fold_num} train val acc.png', dpi = 200, bbox_inches='tight')
    plt.clf()
    
    # Loss
    plt.rcParams["figure.figsize"] = (10,5)
    plt.plot(train_val_acc_loss['train_loss'], label = 'Training loss', color = 'wheat')
    plt.plot(train_val_acc_loss['valid_loss'], label = 'Validation loss', color = 'lawngreen')
    plt.title(f'Best model Loss')
    plt.ylabel('Loss')
    plt.xlabel('Training epoch')
    plt.grid()
    plt.legend(loc=1)   # top right
    plt.savefig(f'{images_dir}/Fold {fold_num} train val loss.png', dpi = 200, bbox_inches='tight')
    plt.clf()

def plot_cm(y_pred, y_test, fold_num, images_dir):
    cm = confusion_matrix(y_test, y_pred)
    cm_display = ConfusionMatrixDisplay(confusion_matrix = cm, display_labels = ['Control (0)', 'Insomnia (1)']).plot(cmap = 'Blues')
    plt.title(f'Confusion Matrix Fold {fold_num}')
    plt.savefig(f'{images_dir}/CM Fold {fold_num}.png', dpi = 100, bbox_inches='tight')
    plt.clf()

    cm_norm = confusion_matrix(y_test, y_pred, normalize = 'true')
    cm_display = ConfusionMatrixDisplay(confusion_matrix = cm_norm, display_labels = ['Control (0)', 'Insomnia (1)']).plot(cmap = 'Blues')
    plt.title(f'Confusion Matrix Normalized Fold {fold_num}')
    plt.savefig(f'{images_dir}/CM Normalized Fold {fold_num}.png', dpi = 100, bbox_inches='tight')
    plt.clf()
    return cm

'''
OTHERS
'''
def scheduler(epoch, lr):
  if epoch < 10:
    return lr
  else:
    return lr * tf.math.exp(-0.1)



# def pipe_image(X, pipe, image_size):
#     X = X.reshape((X.shape[0], X.shape[1] * X.shape[2]))    # Flatten the 224x224 array into a 50176 long array to pass into MinMaxScaler and StandardScaler
#     X = pipe.fit_transform(X)
#     X = X.reshape((X.shape[0], image_size, image_size, 1))    # Reconvert 50176 array into 224x224 array
#     return X

# def save_fold_info(i, hyp_results, performance_metrics, spreadsheet_file):
# def save_validation_results(cv_results, n_splits, spreadsheet_file):
#     book = pxl.load_workbook(spreadsheet_file)
#     writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

#     offset = 0
#     hyperparams = list(cv_results['params'][0].keys())
#     columns1 = ['split'] + hyperparams + ['acc']
#     data1 = []
#     for i in range(n_splits - 1):
#         combos = len(cv_results['params'])
#         #df_fold = pd.DataFrame(data = [cv_results['params']])
#         for params, test_score in zip(cv_results['params'], cv_results[f'split{i}_test_score']):
#             data1.append([i] + list(params.values()) + [test_score])
#     df_cross_val = pd.DataFrame(data = data1, columns = columns1)
#     df_cross_val.to_excel(writer, sheet_name = f'cross_val_results', startrow = 1, startcol = 0,index = False)
#     offset += len(df_cross_val) + 4

#     columns2 = ['combination'] + hyperparams + ['acc_mean'] + ['std_acc_mean']
#     data2 = []
#     for i, (params, mean_test_score, std_test_score) in enumerate(zip(cv_results['params'], cv_results['mean_test_score'], cv_results['std_test_score'])):
#         data2.append([i] + list(params.values()) + [mean_test_score] + [std_test_score])
#     df_cross_val_mean = pd.DataFrame(data = data2, columns = columns2)
#     df_cross_val_mean.to_excel(writer, sheet_name = f'cross_val_results', startrow = offset, startcol = 0, index = False)
#     writer.save()


# def save_test_results(performance_metrics, best_hyperparams, spreadsheet_file):
#     book = pxl.load_workbook(spreadsheet_file)
#     writer = pd.ExcelWriter(spreadsheet_file, engine='openpyxl') 
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 

#     # Best hyperparameters
#     #columns = list(best_hyperparams.keys()) + list(performance_metrics.keys())
#     data = {**best_hyperparams, **performance_metrics}
#     df_test = pd.DataFrame(data = data, index = [0])
#     df_test.to_excel(writer, sheet_name = f'best_model_results', startrow = 1, startcol = 0, index = False)

#     writer.save()



# def plot_fold_test(y_pred, y_test, i, images_dir):

#     plt.plot(y_pred, label = 'pred', linestyle = 'None', markersize = 1.0, marker = '.')
#     plt.plot(y_test, label = 'test')
#     plt.title(f'Fold {i} Test vs predicted')
#     plt.ylabel('Control (0), Insomnia (1)')
#     plt.xlabel('Test epoch')
#     plt.legend()
#     plt.rcParams["figure.figsize"] = (10,5)
#     plt.savefig(f'{images_dir}/Fold {i} Test vs predicted.png', dpi = 200)
#     plt.clf()


# def augment(X_train, label):
#     numEpochDataPoints
#     X_train = GaussianNoise(stddev = std, input_shape = (numEpochDataPoints, 1)),

# def get_train_test(X, y, groups, cv):
    # train_index, test_index = next(cv.split(X, y, groups))
    # X_train = X[train_index].astype('float32')
    # y_train = y[train_index]
    # X_test = X[test_index].astype('float32')
    # y_test = y[test_index]
    # groups_train = groups[train_index]
    # groups_test = groups[test_index]
    # unique_train_groups = np.unique(groups_train)   # Groups only needed for GroupKFold
    # unique_test_groups = np.unique(groups_test)
    # unique, counts = np.unique(y_train, return_counts = True)
    # train_info = dict(zip(unique, counts))
    # unique, counts = np.unique(y_test, return_counts = True)
    # test_info = dict(zip(unique, counts))
    # info = {
    #     'Train groups': str(unique_train_groups),   # To fit list into df
    #     'Test groups': str(unique_test_groups),
    #     'X_train.shape': str(X_train.shape),
    #     'y_train.shape': str(y_train.shape),
    #     'X_test.shape': str(X_test.shape),
    #     'y_test.shape': str(y_test.shape),
    #     'Train class count': str(train_info),
    #     'Test class count': str(test_info)
    # }
    # return X_train, y_train, X_test, y_test, groups_train, groups_test, info

# def plot_train_val_acc_loss(val_acc, val_loss, train_acc, train_loss, images_dir):
    # Accuracy
    # plt.rcParams["figure.figsize"] = (10,5)
    # plt.plot(train_acc, label = 'Training accuracy', color = 'darkorange')
    # plt.plot(val_acc, label = 'Validation accuracy', color = 'darkgreen')
    # plt.title(f'Best model Accuracy')
    # plt.ylabel('Accuracy')
    # plt.xlabel('Training epoch')
    # plt.yticks(np.arange(0, 1, 0.05))
    # plt.grid()
    # plt.legend(bbox_to_anchor = (1, 1))
    # plt.savefig(f'{images_dir}/Best model train val plot.png', dpi = 200)
    # plt.clf()
    
    # # Loss
    # plt.rcParams["figure.figsize"] = (10,5)
    # plt.plot(train_loss, label = 'Training loss', color = 'wheat')
    # plt.plot(val_loss, label = 'Validation loss', color = 'lawngreen')
    # plt.title(f'Best model Loss')
    # plt.ylabel('Loss')
    # plt.xlabel('Training epoch')
    # plt.grid()
    # plt.legend(bbox_to_anchor = (1, 1))
    # plt.savefig(f'{images_dir}/Best model train val loss.png', dpi = 200)

    # def get_excel_writer(spreadsheet_file):
#     return pd.ExcelWriter(spreadsheet_file, engine = 'xlsxwriter')